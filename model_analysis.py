#####################################
#                                   #
#   For Histogram Analysis Code     #
#                                   #
#####################################

import torch.nn as nn
import torch.nn.functional as F
import torch

import matplotlib.pyplot as plt
from PIL import Image
import numpy as np
import math 
import os
from torchvision.utils import save_image
import seaborn as sns
import openpyxl
from openpyxl.utils import get_column_letter

class ResidualBlockBottleNeck(nn.Module):
    def __init__(self, in_features, idx):
        super(ResidualBlockBottleNeck, self).__init__()

        self.idx = idx

        conv_block = [
            nn.Conv2d(in_features, in_features // 4 , 1),
            nn.GroupNorm(in_features // 16, in_features// 4),
            nn.ReLU(inplace=True),
            nn.ReflectionPad2d(1),
            nn.Conv2d(in_features // 4, in_features // 4, 3),
            nn.GroupNorm(in_features // 16, in_features // 4),
            nn.ReLU(inplace=True),
            nn.Conv2d(in_features // 4, in_features, 1),
            nn.GroupNorm(in_features // 4, in_features),
            nn.ReLU(inplace=True),
        ]
        self.conv_block = nn.Sequential(*conv_block)

        self.conv1x1_1 = nn.Conv2d(in_features, in_features // 4 , 1)
        self.gn1 = nn.GroupNorm(in_features // 16, in_features// 4)
        self.conv3x3 = nn.Conv2d(in_features // 4, in_features // 4, 3)
        self.gn2 = nn.GroupNorm(in_features // 16, in_features // 4)
        self.conv1x1_2 = nn.Conv2d(in_features // 4, in_features, 1)
        self.gn3 = nn.GroupNorm(in_features // 4, in_features)
        
        self.relu = nn.ReLU(inplace=True)
        self.reflectionPad = nn.ReflectionPad2d(1)

    def forward(self, x):
        idx = self.idx 

        out = self.conv1x1_1(x)
        histogram(out, "({}) post_conv1x1_residualBlock".format(idx * 10 + 12))
        out = self.gn1(out)
        histogram(out, "({}) post_groupNorm_residualBlock".format(idx * 10 + 13))
        out = self.relu(out)
        histogram(out, "({}) post_relu_residualBlock".format(idx * 10 + 14))

        out = self.reflectionPad(out)

        out = self.conv3x3(out)
        histogram(out, "({}) post_conv3x3_residualBlock".format(idx * 10 + 15))
        out = self.gn2(out)
        histogram(out, "({}) post_groupNorm_residualBlock".format(idx * 10 + 16))
        out = self.relu(out)
        histogram(out, "({}) post_relu_residualBlock".format(idx * 10 + 17))

        out = self.conv1x1_2(out)
        histogram(out, "({}) post_conv1x1_2_residualBlock".format(idx * 10 + 18))
        out = self.gn3(out)
        histogram(out, "({}) post_groupNorm_residualBlock".format(idx * 10 + 19))
        out = self.relu(out)
        histogram(out, "({}) post_relu_residualBlock".format(idx * 10 + 20))

        out = out + x
        histogram(out, "({}) post_residualBlock".format(idx * 10 + 21))

        return out
        # return x + self.conv_block(x)

class ResidualBlock(nn.Module):
    def __init__(self, in_features):
        super(ResidualBlock, self).__init__()

        conv_block = [
            nn.ReflectionPad2d(1),
            nn.Conv2d(in_features, in_features, 3),
            nn.GroupNorm(in_features // 4, in_features),
            nn.ReLU(inplace=True),

            nn.ReflectionPad2d(1),
            nn.Conv2d(in_features, in_features, 3),
            nn.GroupNorm(in_features // 4, in_features),
            nn.ReLU(inplace=True),
        ]
        self.conv_block = nn.Sequential(*conv_block)

    def forward(self, x):
        return x + self.conv_block(x)

analysis_flag = True
global_name = ""
# GroupNorm&UNet&BottleNeck
class Net1(nn.Module):
    def __init__(self, input_nc, output_nc, n_residual_block = 16):
        
        super(Net1, self).__init__()
        self.epoch = 0 
        
        in_features = 64
        out_features = in_features
        # Init Convolution block
        init_block = [
            nn.ReflectionPad2d(3),
            nn.Conv2d(3, in_features, 7),
            nn.GroupNorm(in_features // 4, in_features),
            # nn.Tanh()
        ]
        self.init_block = nn.Sequential(*init_block)
        
        self.init_conv = nn.Conv2d(3, in_features, 7)
        self.init_gn = nn.GroupNorm(in_features // 4, in_features)

        # Downsampling
        downSampleing1 = [
            nn.Conv2d(in_features, out_features * 2, 3, stride=2, padding=1),
            nn.GroupNorm(out_features // 2, out_features * 2),
            # nn.ReLU(inplace=True)
        ]
        downSampleing2 = [
            nn.Conv2d(in_features * 2, out_features * 4, 3, stride=2, padding=1),
            nn.GroupNorm(out_features, out_features * 4),
            # nn.ReLU(inplace=True)
        ]

        self.downSampleing1 = nn.Sequential(*downSampleing1)
        self.downSampleing1_conv = nn.Conv2d(in_features, out_features * 2, 3, stride=2, padding=1)
        self.downSampleing1_gn = nn.GroupNorm(out_features // 2, out_features * 2)

        self.downSampleing2 = nn.Sequential(*downSampleing2)
        self.downSampleing2_conv = nn.Conv2d(in_features * 2, out_features * 4, 3, stride=2, padding=1)
        self.downSampleing2_gn = nn.GroupNorm(out_features, out_features * 4)

        # Residual Blocks
        residualBlock3_list = []
        for idx in range(n_residual_block):
            residualBlock3_list += [ResidualBlockBottleNeck(in_features * 4, idx)]
            # residualBlock3_list += [ResidualBlock(in_features * 4)]


        self.residualBlock3 = nn.Sequential(*residualBlock3_list)
       
        # Upsampling
        upSampling1 = [
            nn.ConvTranspose2d(in_features * 8, out_features * 2, 3, stride=2, padding=1, output_padding=1),
            nn.GroupNorm(out_features // 2, out_features * 2),
            # nn.ReLU(inplace=True)
            ]           
        upSampling2 = [
            nn.ConvTranspose2d(in_features * 4, out_features, 3, stride=2, padding=1, output_padding=1),
            nn.GroupNorm(out_features // 4, out_features),
            # nn.ReLU(inplace=True)
            ]
        self.upSampling1 = nn.Sequential(*upSampling1)
        self.upSampling1_deconv = nn.ConvTranspose2d(in_features * 8, out_features * 2, 3, stride=2, padding=1, output_padding=1)
        self.upSampling1_gn = nn.GroupNorm(out_features // 2, out_features * 2)

        self.upSampling2 = nn.Sequential(*upSampling2)
        self.upSampling2_deconv = nn.ConvTranspose2d(in_features * 4, out_features, 3, stride=2, padding=1, output_padding=1)
        self.upSampling2_gn = nn.GroupNorm(out_features // 4, out_features)

        # Output layer
        out_block = [
            nn.ReflectionPad2d(3),
            nn.Conv2d(in_features, 3, 7),
            # nn.Tanh()
            ]
        self.out_block = nn.Sequential(*out_block)
        self.out_block_conv = nn.Conv2d(in_features, 3, 7)

        self.tanh = nn.Tanh()
        self.relu = nn.ReLU()
        self.reflectionPad = nn.ReflectionPad2d(3)

    def forward(self, x, name="", analysis=True):

        global analysis_flag
        analysis_flag = analysis

        global global_name
        global_name = name

        histogram(x[0], "(1) input")
        idx = 0
        # for layer in x[0]:
        #     idx = idx + 1
        #     histogram(layer, "input_{}".format(idx), "input")

        # init_block = self.init_block(x)
        # histogram(init_block, "(2) pre_tanh_init_block")
        # init_block = self.tanh(init_block)

        init_block = self.reflectionPad(x)
        histogram(init_block, "(2) pre_conv_init_block")
        # print(self.init_conv.parameters())
        init_block = self.init_conv(init_block)
        histogram(init_block, "(3) post_conv_init_block")

        init_block = self.init_gn(init_block)
        histogram(init_block, "(4) post_groupNorm_init_block")

        init_block = self.tanh(init_block)
        histogram(init_block, "(5) post_relu_init_block")


        idx = 0
        # histogram(init_block, "(3) post_tanh_init_block")
        # for layer in init_block[0]:
        #     idx = idx + 1
        #     histogram(layer, "init_block_{}".format(idx), "init_block")

            # print(idx)
        #     writer.add_histogram("post_init_block", layer.clone().cpu().data.numpy(), idx)

        # downSampleing1 = self.downSampleing1(init_block)

        # histogram(downSampleing1, "(4) pre_relu_downSampling1")
        # downSampleing1 = self.relu(downSampleing1)
        # histogram(downSampleing1, "(5) post_relu_downSampling1")

        downSampleing1 = self.downSampleing1_conv(init_block)
        histogram(downSampleing1, "(6) post_conv_downSampling1")
        downSampleing1 = self.downSampleing1_gn(downSampleing1)
        histogram(downSampleing1, "(7) post_groupNorm_downSampling1")
        downSampleing1 = self.relu(downSampleing1)
        histogram(downSampleing1, "(8) post_relu_downSampling1")

        # print(downSampleing1.shape)
        idx = 0

        # for layer in downSampleing1[0]:
        #     idx = idx + 1
        #     histogram(layer, "downSampling1_{}".format(idx), "downSampling1")

        #     writer.add_histogram("post_encoder_block1", layer.clone().cpu().data.numpy(), idx)

        # downSampleing2 = self.downSampleing2(downSampleing1)

        # histogram(downSampleing2, "(6) pre_relu_downSampling2")
        # downSampleing2 = self.relu(downSampleing2)
        # histogram(downSampleing2, "(7) post_relu_downSampling2")

        downSampleing2 = self.downSampleing2_conv(downSampleing1)
        histogram(downSampleing2, "(9) post_conv_downSampling2")
        downSampleing2 = self.downSampleing2_gn(downSampleing2)
        histogram(init_block, "(10) post_groupNorm_downSampling2")
        downSampleing2 = self.relu(downSampleing2)
        histogram(downSampleing1, "(11) post_relu_downSampling2")

        # idx = 0
        # for layer in downSampleing2[0]:
        #     idx = idx + 1
        #     histogram(layer, "downSampling2_{}".format(idx), "downSampling2")
            
            # writer.add_histogram("post_encoder_block2", layer.clone().cpu().data.numpy(), idx)

        residualBlock3 = self.residualBlock3(downSampleing2)

        # idx = 0
        # for layer in residualBlock3[0]:
        #     idx = idx + 1
        #     histogram(layer, "residualBlock3_{}".format(idx), "residualBlock3")

        #     writer.add_histogram("post_residualmodule", layer.clone().cpu().data.numpy(), idx)

        # upSampling1 = self.upSampling1(torch.cat([downSampleing2, residualBlock3], 1))          # 128
        # histogram(upSampling1, "(9) pre_relu_upSampling1")
        # upSampling1 = self.relu(upSampling1)
        # histogram(upSampling1, "(10) post_rele_upSampling1")
        # 172
        concat1 = torch.cat([downSampleing2, residualBlock3], 1)
        histogram(concat1, "(172) pre_deconv_upSampling1(concat)")
        upSampling1 = self.upSampling1_deconv(concat1)          # 128
        histogram(upSampling1, "(173) post_deconv_upSampling1")
        upSampling1 = self.upSampling1_gn(upSampling1)
        histogram(upSampling1, "(174) post_groupNorm_upSampling1")
        upSampling1 = self.relu(upSampling1)
        histogram(upSampling1, "(175) post_relu_upSampling1")

        # idx = 0
        # for layer in upSampling1[0]:
        #     idx = idx + 1
        #     histogram(layer, "upSampling1_{}".format(idx), "upSampling1")
        #     writer.add_histogram("post_decoder_block1", layer.clone().cpu().data.numpy(), idx)

        # upSampling2 = self.upSampling2(torch.cat([upSampling1, downSampleing1], 1))          # 64
        # histogram(upSampling2, "(11) pre_relu_upSampling2")
        # upSampling2 = self.relu(upSampling2)
        # histogram(upSampling2, "(12) post_relu_upSampling2")

        concat2 = torch.cat([upSampling1, downSampleing1], 1)
        histogram(concat2, "(176) pre_deconv_upSampling2(concat)")

        upSampling2 = self.upSampling2_deconv(concat2)          # 128
        histogram(upSampling2, "(177) post_deconv_upSampling2")
        upSampling2 = self.upSampling2_gn(upSampling2)
        histogram(upSampling2, "(178) post_groupNorm_upSampling2")
        upSampling2 = self.relu(upSampling2)
        histogram(upSampling2, "(179) post_relu_upSampling2")

        # idx = 0
        # for layer in upSampling2[0]:
        #     idx = idx + 1
        #     histogram(layer, "upSampling2_{}".format(idx), "upSampling2")
            # print(layer.shape)
        #     writer.add_histogram("post_decoder_block2", layer.clone().cpu().data.numpy(), idx)

        # out_block = self.out_block(upSampling2)
        # histogram(out_block, "(13) pre_tanh_out_block")
        # out_block = self.tanh(out_block)
        # histogram(out_block, "(14) post_tanh_out_block")

        out_block = self.reflectionPad(upSampling2)
        out_block = self.out_block_conv(out_block)
        histogram(out_block, "(180) post_conv_out_block")
        # out_block = self.out_block_gn(out_block)
        # histogram(out_block, "(181) post_groupNorm_out_block")
        out_block = self.tanh(out_block)
        histogram(out_block, "(181) post_tanh_out_block")

        # idx = 0
        # for layer in out_block[0]:
        #     idx = idx + 1
        #     histogram(layer, "out_block_{}".format(idx), "out_block")

        #     writer.add_histogram("post_out_block", layer.clone().cpu().data.numpy(), idx)

        out = x + out_block

        # J = I + R
        histogram(out, "(182) output")

        # hist1 = torch.histc(x, bins=100, min=-1, max=1).clone().cpu().data.numpy()
        # hist2 = torch.histc(out_block, bins=100, min=-1, max=1).clone().cpu().data.numpy()
        # fig = plt.figure()
        # plt.bar(np.arange(100), hist1, color='b')
        # plt.bar(np.arange(100), hist2, color='r', bottom=hist1)
        # plt.xticks(np.arange(100, step=10), ["%.2f" %x for x in np.arange(-1, 1, 0.2)], rotation=30)
        # plt.savefig("./histogram/(46) x_out_block.png", dpi=300)
        # plt.close()
                
        return out


def histogram(tensor, name, dir=""):
    global analysis_flag
    if not analysis_flag: return 
    global global_name

    filename = global_name.split(".jpg")[0]
    # try:
    #     wb = openpyxl.load_workbook("./histogram/{}_{}.xlsx".format(filename))
    # except:
    wb = openpyxl.Workbook()


    features = torch.squeeze(tensor)
    # print(feautres.shape)

    # path = "./histogram/{}/{}".format(global_name, dir)
    # if not(os.path.isdir(path)):
    #     os.makedirs(os.path.join(path))
    
    sheet = wb.create_sheet(name)

    row = 1
    for idx, feature in enumerate(features):
    #     MAX = torch.max(feature).cpu().data
    #     MIN = torch.min(feature).cpu().data
    #     print("MIN: %.4f, " % MIN + "MAX: %.4f" % MAX)
        # print(global_name)
        # print(name)
        sheet.append([''])
        col = 1
        print(feature.shape)
        for r, k in enumerate(feature):
            # cols = []
            for c, l in enumerate(k):
                # column_variable =  get_column_letter(col+c+1)
                # print(column_variable)
                sheet.cell((row+r), (col+c), l.item())
                
                # print(type(l.item()))
                # cols.append(l.item())
                # line += "{} ".format(l)
            # print(list(k.size())[0])
        row = row + list(feature.size())[0] + 1
    col = col + list(k.size())[0] + 1
    
    wb.save("./histogram/{}_{}.xlsx".format(filename, name))
        # path = "./histogram/{}/{}".format(global_name, name)
        # if not(os.path.isdir(path)):
        #     os.makedirs(os.path.join(path))
    
        # sns.heatmap(feature.cpu(), cmap="YlGnBu")
        # heatmap_file = "{} {}.png".format(name, idx + 1)
        # plt.title(heatmap_file)
        # # plt.show()
        # plt.savefig(path+"/{}.png".format(heatmap_file), dpi=300)
        # plt.close()

        # print(feature.shape)

    # heatmap = tensor
    # heatmap = np.uint8(255*heatmap)


    # f = open(path + "/{}.txt".format(name), 'w')
    
    MAX = torch.max(tensor).cpu().data
    MIN = torch.min(tensor).cpu().data
    print("MIN: %.4f, " % MIN + "MAX: %.4f" % MAX)

    # BINS = 100
    
    # hist = torch.histc(tensor, bins=BINS, min=MIN, max=MAX).clone().cpu().data.numpy()

    # line = "{}\n".format(name)
    # for _, i in enumerate(hist):
    #     line = line + "{}, ".format(i)
    # line = line + "\nMIN: {}, MAX: {} BINS: {}".format(MIN, MAX, BINS)
    # f.write(line)

    # fig = plt.figure()
    # rng = np.arange(len(hist))

    # plt.bar(rng, hist)

    # if MAX - MIN == 0:
    #     temp = MIN.clone().cpu().data.numpy()
    #     plt.xticks(np.arange(len(hist), step=len(hist)/(BINS//10)), ["","","","", temp - 1, temp, temp + 1,"","",""])
    # else:
    #     plt.xticks(np.arange(len(hist), step=len(hist)/(BINS//10)), ["%.2f" %x for x in np.arange(MIN, MAX, (MAX-MIN)/(BINS//10))], rotation=30)

    # plt.savefig(path+"/{}.png".format(name), dpi=300)
    # plt.close()
