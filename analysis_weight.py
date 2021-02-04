
import torch
from PIL import Image
from torch.autograd import Variable
from torchvision.transforms import transforms
from torchvision.utils import save_image

from model_analysis import Net1

import openpyxl
from openpyxl.utils import get_column_letter
TEST_MODE = True if torch.cuda.is_available() else False

def Analysis_weight(model, file_name):
    wb = openpyxl.Workbook()
    
    for name, parameter in model.named_parameters():
        if not parameter.requires_grad: 
            continue
        param = parameter.numel()
        temp_name = name.split(".")
        if temp_name[0] == "init_block" or temp_name[0] == "downSampleing1" or temp_name[0] == "downSampleing2" or temp_name[0] == "upSampling1" or temp_name[0] == "upSampling2" or temp_name[0] == "out_block" or len(temp_name) >= 5:
            continue
        if temp_name[-1] == "bias": continue
        if temp_name[-2] == "gn" or temp_name[-2] == "gn1" or temp_name[-2] == "gn2" or temp_name[-2] == "gn3": continue
        if temp_name[0] == "init_gn" or temp_name[0] == "downSampleing1_gn" or temp_name[0] == "downSampleing2_gn" or temp_name[0] == "upSampling1_gn" or temp_name[0] == "upSampling2_gn": continue
        
        # print(parameter.shape)
        name = ""
        if len(temp_name) > 2:
            name = "{}_{}_{}".format(temp_name[0], temp_name[1], temp_name[2])
        else: 
            name = "{}".format(temp_name[0])
        print(name)
        sheet = wb.create_sheet(name)
        row = 1
        for idx_in, i in enumerate(parameter):
            sheet.append([''])
            col = 1
            # filters = []
            for idx_out, j in enumerate(i):

                # if len(temp_name) > 2:
                #     name = "{}_{}_{}_{}.txt".format(temp_name[0],temp_name[1], temp_name[2], idx_out)
                # else: 
                #     name = "{}_{}.txt".format(temp_name[0], idx_out)
                # rows = []
                for r, k in enumerate(j):
                    # cols = []
                    for c, l in enumerate(k):
                        # column_variable =  get_column_letter(col+c+1)
                        # print(column_variable)
                        sheet.cell((row+r), (col+c), l.item())
                        
                        # print(type(l.item()))
                        # cols.append(l.item())
                        # line += "{} ".format(l)
                    # print(list(k.size())[0])
                col = col + list(k.size())[0] + 1
            row = row + list(j.size())[0] + 1
                    # rows.append(cols)
                    # sheet.append(line)
                    # line = []                
                # filters.append(rows)
                    # line += "\n"
            # sheet.append(filters)

            print(col, row)  
            #     line += "\n"
            # line += "\n"
                


        # params = params + param
        # print(name, param)
    # print(params)

    wb.save("./weight/{}.xlsx".format(file_name))

# train_info = "LoG5x5_analysis"
# for epoch in range(100):
#     if train_info != "":
#         checkpoint_dir = "./checkpoints/cp_{}".format(train_info)
#     else:
#         checkpoint_dir = "./checkpoints/"

#     net = Net1(3, 3)
#     model = net.eval()
#     model.load_state_dict(torch.load(checkpoint_dir + '/checkpoint_{}.pth'.format(epoch+1)))
#     model.cuda()
#     Analysis_weight(model)
